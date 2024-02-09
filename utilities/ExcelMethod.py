import openpyxl

def numRows(file_name,sheet_name):
    Excel_file = openpyxl.load_workbook(file_name) # loading the file
    Sheet = Excel_file[sheet_name] # loading the sheet
    return Sheet.max_row # returning the maximum number of rows in the sheet

def readData(file_name, sheet_name, row_num, col_num):
    Excel_file = openpyxl.load_workbook(file_name)
    Sheet =Excel_file[sheet_name]
    return Sheet.cell(row=row_num, column=col_num).value # returning the value of given row column


def writeData(file_name,sheet_name,row_num,col_num,data):
    Excel_file = openpyxl.load_workbook(file_name)
    Sheet = Excel_file[sheet_name]
    Sheet.cell(row=row_num, column = col_num).value = data # writing value at given row and column (cell)
    Excel_file.save(file_name)



